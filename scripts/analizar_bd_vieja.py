import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Usuario\Downloads\Gestion Taller Muñoz .xlsx", data_only=True)
ws = wb["Base de datos vieja"]

headers = [cell for cell in next(ws.iter_rows(values_only=True))]
print("=== BASE DE DATOS VIEJA STATS ===")
total_rows = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(row):
        total_rows += 1

print(f"Total non-empty rows in 'Base de datos vieja': {total_rows}")
