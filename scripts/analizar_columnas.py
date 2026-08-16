import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Usuario\Downloads\Gestion Taller Muñoz .xlsx", data_only=True)
ws = wb["Base de datos"]

headers = [cell for cell in next(ws.iter_rows(values_only=True))]
print("=== BASE DE DATOS COLUMNS & STATS ===")
col_stats = {h: 0 for h in headers if h}
total_rows = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    total_rows += 1
    for i, h in enumerate(headers):
        if h and i < len(row) and row[i] is not None:
            col_stats[h] += 1

print(f"Total rows: {total_rows}")
for h, count in col_stats.items():
    pct = (count / total_rows) * 100 if total_rows > 0 else 0
    print(f"  - Column '{h.strip()}': {count} non-empty values ({pct:.1f}%)")
