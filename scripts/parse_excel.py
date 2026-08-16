import openpyxl
import json
import re
from datetime import datetime

EXCEL_PATH = r"C:\Users\Usuario\Downloads\Gestion Taller Muñoz .xlsx"
OUTPUT_JSON = r"scripts\taller_munoz_data.json"

def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def clean_num(val):
    if val is None:
        return 0
    try:
        return float(val)
    except:
        return 0

def clean_date(val):
    if isinstance(val, datetime):
        return val.isoformat()
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def parse():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # 1. Base de datos
    ws_bd = wb['Base de datos']
    headers_bd = [str(c).strip() if c is not None else f'col_{i}' for i, c in enumerate(next(ws_bd.iter_rows(values_only=True)))]
    
    ordenes = []
    for row in ws_bd.iter_rows(min_row=2, values_only=True):
        d = {headers_bd[i]: row[i] for i in range(len(headers_bd)) if i < len(row)}
        
        patente = clean_str(d.get('Patente'))
        vehiculo = clean_str(d.get('Vehiculo'))
        cliente = clean_str(d.get('Nombre del cliente'))
        ot_num = clean_str(d.get('Orden de trabajo'))
        
        if not patente and not vehiculo and not cliente and not ot_num:
            continue
            
        km = clean_num(d.get('Kilometraje'))
        mano_obra = clean_num(d.get('Cobro mano de obra'))
        repuestos = clean_num(d.get('Cobro repuesto'))
        fecha_ingreso = clean_date(d.get('paleta ')) or clean_date(d.get('paleta')) or datetime.now().isoformat()
        fecha_fin = clean_date(d.get('Fecha finalizado'))
        estado_raw = clean_str(d.get('Estado'))
        
        anomalias = []
        for a_key in ['Anomalia del cliente ', 'Anomalia del cliente 2', 'Anomalia del cliente 3', 'Anomalia del cliente 4']:
            a_val = clean_str(d.get(a_key))
            if a_val:
                anomalias.append(a_val)
                
        tercerizados = clean_str(d.get('Tercerizados'))
        
        ordenes.append({
            'ot_num': ot_num,
            'patente': patente,
            'vehiculo': vehiculo,
            'cliente': cliente,
            'km': int(km) if km else None,
            'mano_obra': mano_obra,
            'repuestos': repuestos,
            'total': mano_obra + repuestos,
            'fecha_ingreso': fecha_ingreso,
            'fecha_fin': fecha_fin,
            'estado': estado_raw,
            'anomalias': anomalias,
            'tercerizados': tercerizados,
        })
        
    print(f"Parsed {len(ordenes)} ordenes from Base de datos")
    
    # 2. Turnos
    ws_turnos = wb['Turnos']
    turnos = []
    for row in ws_turnos.iter_rows(values_only=True):
        if not any(row):
            continue
        fecha = clean_date(row[0])
        turno_horario = clean_str(row[1]) if len(row) > 1 else None
        cliente = clean_str(row[2]) if len(row) > 2 else None
        auto = clean_str(row[3]) if len(row) > 3 else None
        
        if fecha or cliente or auto:
            turnos.append({
                'fecha': fecha,
                'horario': turno_horario,
                'cliente': cliente,
                'vehiculo': auto
            })
            
    print(f"Parsed {len(turnos)} turnos")

    data = {
        'ordenes': ordenes,
        'turnos': turnos
    }
    
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        
    print(f"Saved to {OUTPUT_JSON}")

if __name__ == '__main__':
    parse()
